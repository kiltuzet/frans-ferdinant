import sys
import os
import hashlib
import shutil
import datetime
import sqlite3
from PyQt5.QtWidgets import (
    QApplication, QWidget, QTabWidget, QVBoxLayout, QPushButton, QLabel,
    QLineEdit, QHBoxLayout, QMessageBox, QTableWidget, QTableWidgetItem,
    QDateEdit
)
from PyQt5.QtCore import QDate, QTimer
import openpyxl

# Хеширование пароля SHA256
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

class LoginWindow(QWidget):
    def __init__(self, on_success):
        super().__init__()
        self.on_success = on_success
        self.setWindowTitle("Вход")
        self.resize(300, 150)

        self.conn = sqlite3.connect("buspark_filled.db")
        self.cursor = self.conn.cursor()
        self.initDB()

        layout = QVBoxLayout()
        layout.addWidget(QLabel("Имя пользователя:"))
        self.username_input = QLineEdit()
        layout.addWidget(self.username_input)

        layout.addWidget(QLabel("Пароль:"))
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password_input)

        login_btn = QPushButton("Войти")
        login_btn.clicked.connect(self.check_login)
        layout.addWidget(login_btn)

        self.setLayout(layout)

    def initDB(self):
        # Таблица пользователей с ролью
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE,
                password TEXT,
                role TEXT DEFAULT 'user'
            )
        ''')
        self.conn.commit()

        # Создаем админа, если не существует
        hashed_admin = hash_password("admin")
        self.cursor.execute("INSERT OR IGNORE INTO users (username, password, role) VALUES (?, ?, ?)",
                            ("admin", hashed_admin, "admin"))
        self.conn.commit()

    def check_login(self):
        username = self.username_input.text()
        password = self.password_input.text()
        hashed_pass = hash_password(password)

        try:
            self.cursor.execute("SELECT id, role FROM users WHERE username=? AND password=?", (username, hashed_pass))
            result = self.cursor.fetchone()
            if result:
                user_id, role = result
                self.on_success(user_id, role)
                self.close()
            else:
                QMessageBox.warning(self, "Ошибка", "Неверный логин или пароль")
        except sqlite3.Error as e:
            QMessageBox.critical(self, "Ошибка базы данных", f"Ошибка:\n{e}")

class BusParkApp(QWidget):
    def __init__(self, user_id, user_role):
        super().__init__()
        self.user_id = user_id
        self.user_role = user_role

        self.setWindowTitle("Автоматизированная система Автобусный парк")
        self.resize(900, 600)

        self.conn = sqlite3.connect("buspark_filled.db")
        self.cursor = self.conn.cursor()

        self.initDB()
        self.initUI()
        #self.check_maintenance_notifications()

        # Таймер для ежедневной проверки ТО
        #timer = QTimer(self)
        #timer.timeout.connect(self.check_maintenance_notifications)
        #timer.start(24 * 3600 * 1000)  # 24 часа

    def initDB(self):
        # Таблицы
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS buses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model TEXT,
                number TEXT UNIQUE,
                year INTEGER,
                status TEXT
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS drivers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                license_number TEXT UNIQUE,
                phone TEXT
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS routes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                number TEXT UNIQUE,
                description TEXT
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS maintenance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bus_id INTEGER,
                date TEXT,
                type TEXT,
                FOREIGN KEY(bus_id) REFERENCES buses(id)
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                action TEXT,
                timestamp TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        self.conn.commit()

    def initUI(self):
        layout = QVBoxLayout()

        self.tabs = QTabWidget()

        self.bus_tab = self.createBusTab()
        self.driver_tab = self.createDriverTab()
        self.route_tab = self.createRouteTab()
        self.maint_tab = self.createMaintenanceTab()
        self.reports_tab = self.createReportsTab()

        self.tabs.addTab(self.bus_tab, "Автобусы")
        self.tabs.addTab(self.driver_tab, "Водители")
        self.tabs.addTab(self.route_tab, "Маршруты")
        self.tabs.addTab(self.maint_tab, "ТО")
        self.tabs.addTab(self.reports_tab, "Отчёты")

        layout.addWidget(self.tabs)

        # Кнопка резервного копирования
       # backup_btn = QPushButton("Создать резервную копию БД")
       # backup_btn.clicked.connect(self.backup_database)
        #layout.addWidget(backup_btn)
        print("all ake")
        self.setLayout(layout)

    # --- Логирование действий ---
    def log_action(self, action):
        self.cursor.execute("INSERT INTO user_logs (user_id, action) VALUES (?, ?)", (self.user_id, action))
        self.conn.commit()

    # --- Резервное копирование ---
    def backup_database(self):
        now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = "backups"
        os.makedirs(backup_dir, exist_ok=True)
        backup_path = os.path.join(backup_dir, f"buspark_backup_{now}.db")
        self.conn.close()
        shutil.copy("buspark_filled.db", backup_path)
        self.conn = sqlite3.connect("buspark_filled.db")
        self.cursor = self.conn.cursor()
        QMessageBox.information(self, "Резервное копирование", f"База сохранена в {backup_path}")
        self.log_action(f"Создана резервная копия базы: {backup_path}")

    # --- Создание вкладок ---
    def createBusTab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        # Поля ввода
        hlayout = QHBoxLayout()
        hlayout.addWidget(QLabel("Модель:"))
        self.bus_model = QLineEdit()
        hlayout.addWidget(self.bus_model)

        hlayout.addWidget(QLabel("Номер:"))
        self.bus_number = QLineEdit()
        hlayout.addWidget(self.bus_number)

        hlayout.addWidget(QLabel("Год:"))
        self.bus_year = QLineEdit()
        hlayout.addWidget(self.bus_year)

        hlayout.addWidget(QLabel("Статус:"))
        self.bus_status = QLineEdit()
        hlayout.addWidget(self.bus_status)

        layout.addLayout(hlayout)

        # Кнопки
        btn_layout = QHBoxLayout()
        self.bus_save_btn = QPushButton("Сохранить автобус")
        self.bus_save_btn.clicked.connect(self.saveBus)
        btn_layout.addWidget(self.bus_save_btn)

        self.bus_export_btn = QPushButton("Экспорт в Excel")
        self.bus_export_btn.clicked.connect(self.exportBusesToExcel)
        btn_layout.addWidget(self.bus_export_btn)

        layout.addLayout(btn_layout)

        # Таблица
        self.bus_table = QTableWidget()
        layout.addWidget(self.bus_table)

        tab.setLayout(layout)

        self.loadBuses()

        # Ограничения по ролям
        if self.user_role != "admin":
            self.bus_save_btn.setEnabled(False)

        return tab

    def createDriverTab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        hlayout = QHBoxLayout()
        hlayout.addWidget(QLabel("Имя:"))
        self.driver_name = QLineEdit()
        hlayout.addWidget(self.driver_name)

        hlayout.addWidget(QLabel("Номер лицензии:"))
        self.driver_license = QLineEdit()
        hlayout.addWidget(self.driver_license)

        hlayout.addWidget(QLabel("Телефон:"))
        self.driver_phone = QLineEdit()
        hlayout.addWidget(self.driver_phone)

        layout.addLayout(hlayout)

        btn_layout = QHBoxLayout()
        self.driver_save_btn = QPushButton("Сохранить водителя")
        self.driver_save_btn.clicked.connect(self.saveDriver)
        btn_layout.addWidget(self.driver_save_btn)

        self.driver_export_btn = QPushButton("Экспорт в Excel")
        self.driver_export_btn.clicked.connect(self.exportDriversToExcel)
        btn_layout.addWidget(self.driver_export_btn)

        layout.addLayout(btn_layout)

        self.driver_table = QTableWidget()
        layout.addWidget(self.driver_table)

        tab.setLayout(layout)

        self.loadDrivers()

        if self.user_role != "admin":
            self.driver_save_btn.setEnabled(False)

        return tab

    def createRouteTab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        hlayout = QHBoxLayout()
        hlayout.addWidget(QLabel("Номер маршрута:"))
        self.route_number = QLineEdit()
        hlayout.addWidget(self.route_number)

        hlayout.addWidget(QLabel("Описание:"))
        self.route_desc = QLineEdit()
        hlayout.addWidget(self.route_desc)

        layout.addLayout(hlayout)

        btn_layout = QHBoxLayout()
        self.route_save_btn = QPushButton("Сохранить маршрут")
        self.route_save_btn.clicked.connect(self.saveRoute)
        btn_layout.addWidget(self.route_save_btn)

        self.route_export_btn = QPushButton("Экспорт в Excel")
        self.route_export_btn.clicked.connect(self.exportRoutesToExcel)
        btn_layout.addWidget(self.route_export_btn)

        layout.addLayout(btn_layout)

        self.route_table = QTableWidget()
        layout.addWidget(self.route_table)

        tab.setLayout(layout)

        self.loadRoutes()

        if self.user_role != "admin":
            self.route_save_btn.setEnabled(False)

        return tab

    def createMaintenanceTab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        hlayout = QHBoxLayout()
        hlayout.addWidget(QLabel("ID автобуса:"))
        self.maint_bus_id = QLineEdit()
        hlayout.addWidget(self.maint_bus_id)

        hlayout.addWidget(QLabel("Дата (ГГГГ-ММ-ДД):"))
        self.maint_date = QLineEdit()
        hlayout.addWidget(self.maint_date)

        hlayout.addWidget(QLabel("Тип ТО:"))
        self.maint_type = QLineEdit()
        hlayout.addWidget(self.maint_type)

        layout.addLayout(hlayout)

        btn_layout = QHBoxLayout()
        self.maint_save_btn = QPushButton("Сохранить ТО")
        self.maint_save_btn.clicked.connect(self.saveMaintenance)
        btn_layout.addWidget(self.maint_save_btn)

        self.maint_export_btn = QPushButton("Экспорт в Excel")
        self.maint_export_btn.clicked.connect(self.exportMaintenanceToExcel)
        btn_layout.addWidget(self.maint_export_btn)

        layout.addLayout(btn_layout)

        self.maint_table = QTableWidget()
        layout.addWidget(self.maint_table)

        tab.setLayout(layout)

        self.loadMaintenance()

        if self.user_role != "admin":
            self.maint_save_btn.setEnabled(False)

        return tab

    def createReportsTab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        layout.addWidget(QLabel("Выберите дату ТО:"))
        self.report_date = QDateEdit()
        self.report_date.setCalendarPopup(True)
        self.report_date.setDate(QDate.currentDate())
        layout.addWidget(self.report_date)

        gen_report_btn = QPushButton("Сгенерировать отчёт ТО")
        gen_report_btn.clicked.connect(self.generateMaintenanceReport)
        layout.addWidget(gen_report_btn)

        self.report_table = QTableWidget()
        layout.addWidget(self.report_table)

        tab.setLayout(layout)
        return tab

    # --- Загрузка данных ---
    def loadBuses(self):
        self.cursor.execute("SELECT * FROM buses")
        rows = self.cursor.fetchall()
        self.bus_table.setColumnCount(5)
        self.bus_table.setRowCount(len(rows))
        self.bus_table.setHorizontalHeaderLabels(["ID", "Модель", "Номер", "Год", "Статус"])
        for row_idx, row in enumerate(rows):
            for col_idx, val in enumerate(row):
                self.bus_table.setItem(row_idx, col_idx, QTableWidgetItem(str(val)))

    def loadDrivers(self):
        self.cursor.execute("SELECT * FROM drivers")
        rows = self.cursor.fetchall()
        self.driver_table.setColumnCount(4)
        self.driver_table.setRowCount(len(rows))
        self.driver_table.setHorizontalHeaderLabels(["ID", "Имя", "Номер лицензии", "Телефон"])
        for row_idx, row in enumerate(rows):
            for col_idx, val in enumerate(row):
                self.driver_table.setItem(row_idx, col_idx, QTableWidgetItem(str(val)))

    def loadRoutes(self):
        self.cursor.execute("SELECT * FROM routes")
        rows = self.cursor.fetchall()
        self.route_table.setColumnCount(3)
        self.route_table.setRowCount(len(rows))
        self.route_table.setHorizontalHeaderLabels(["ID", "Номер маршрута", "Описание"])
        for row_idx, row in enumerate(rows):
            for col_idx, val in enumerate(row):
                self.route_table.setItem(row_idx, col_idx, QTableWidgetItem(str(val)))

    def loadMaintenance(self):
        self.cursor.execute("SELECT * FROM maintenance")
        rows = self.cursor.fetchall()
        self.maint_table.setColumnCount(4)
        self.maint_table.setRowCount(len(rows))
        self.maint_table.setHorizontalHeaderLabels(["ID", "ID автобуса", "Дата", "Тип"])
        for row_idx, row in enumerate(rows):
            for col_idx, val in enumerate(row):
                self.maint_table.setItem(row_idx, col_idx, QTableWidgetItem(str(val)))

    # --- Сохранение данных ---
    def saveBus(self):
        model = self.bus_model.text()
        number = self.bus_number.text()
        year = self.bus_year.text()
        status = self.bus_status.text()
        if not (model and number and year and status):
            QMessageBox.warning(self, "Ошибка", "Заполните все поля")
            return
        try:
            year_int = int(year)
        except ValueError:
            QMessageBox.warning(self, "Ошибка", "Год должен быть числом")
            return

        try:
            self.cursor.execute("INSERT INTO buses (model, number, year, status) VALUES (?, ?, ?, ?)",
                                (model, number, year_int, status))
            self.conn.commit()
            self.loadBuses()
            self.log_action(f"Добавлен автобус: модель={model}, номер={number}, год={year}, статус={status}")
            QMessageBox.information(self, "Успех", "Автобус добавлен")
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Ошибка", "Автобус с таким номером уже существует")

    def saveDriver(self):
        name = self.driver_name.text()
        license_num = self.driver_license.text()
        phone = self.driver_phone.text()
        if not (name and license_num and phone):
            QMessageBox.warning(self, "Ошибка", "Заполните все поля")
            return
        try:
            self.cursor.execute("INSERT INTO drivers (name, license_number, phone) VALUES (?, ?, ?)",
                                (name, license_num, phone))
            self.conn.commit()
            self.loadDrivers()
            self.log_action(f"Добавлен водитель: имя={name}, лицензия={license_num}, телефон={phone}")
            QMessageBox.information(self, "Успех", "Водитель добавлен")
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Ошибка", "Водитель с таким номером лицензии уже существует")

    def saveRoute(self):
        number = self.route_number.text()
        desc = self.route_desc.text()
        if not (number and desc):
            QMessageBox.warning(self, "Ошибка", "Заполните все поля")
            return
        try:
            self.cursor.execute("INSERT INTO routes (number, description) VALUES (?, ?)", (number, desc))
            self.conn.commit()
            self.loadRoutes()
            self.log_action(f"Добавлен маршрут: номер={number}, описание={desc}")
            QMessageBox.information(self, "Успех", "Маршрут добавлен")
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Ошибка", "Маршрут с таким номером уже существует")

    def saveMaintenance(self):
        bus_id = self.maint_bus_id.text()
        date = self.maint_date.text()
        mtype = self.maint_type.text()
        if not (bus_id and date and mtype):
            QMessageBox.warning(self, "Ошибка", "Заполните все поля")
            return
        try:
            bus_id_int = int(bus_id)
        except ValueError:
            QMessageBox.warning(self, "Ошибка", "ID автобуса должен быть числом")
            return

        try:
            # Проверяем формат даты
            datetime.datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            QMessageBox.warning(self, "Ошибка", "Дата должна быть в формате ГГГГ-ММ-ДД")
            return

        try:
            self.cursor.execute("INSERT INTO maintenance (bus_id, date, type) VALUES (?, ?, ?)",
                                (bus_id_int, date, mtype))
            self.conn.commit()
            self.loadMaintenance()
            self.log_action(f"Добавлено ТО: автобус_id={bus_id_int}, дата={date}, тип={mtype}")
            QMessageBox.information(self, "Успех", "ТО добавлено")
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Ошибка", "Ошибка при добавлении ТО")

    # --- Экспорт в Excel ---
    def exportBusesToExcel(self):
        self.cursor.execute("SELECT * FROM buses")
        rows = self.cursor.fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Автобусы"
        ws.append(["ID", "Модель", "Номер", "Год", "Статус"])
        for r in rows:
            ws.append(r)
        filename = "buses_export.xlsx"
        wb.save(filename)
        QMessageBox.information(self, "Экспорт", f"Данные автобусов экспортированы в {filename}")
        self.log_action("Экспорт данных автобусов в Excel")

    def exportDriversToExcel(self):
        self.cursor.execute("SELECT * FROM drivers")
        rows = self.cursor.fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Водители"
        ws.append(["ID", "Имя", "Номер лицензии", "Телефон"])
        for r in rows:
            ws.append(r)
        filename = "drivers_export.xlsx"
        wb.save(filename)
        QMessageBox.information(self, "Экспорт", f"Данные водителей экспортированы в {filename}")
        self.log_action("Экспорт данных водителей в Excel")

    def exportRoutesToExcel(self):
        self.cursor.execute("SELECT * FROM routes")
        rows = self.cursor.fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Маршруты"
        ws.append(["ID", "Номер маршрута", "Описание"])
        for r in rows:
            ws.append(r)
        filename = "routes_export.xlsx"
        wb.save(filename)
        QMessageBox.information(self, "Экспорт", f"Данные маршрутов экспортированы в {filename}")
        self.log_action("Экспорт данных маршрутов в Excel")

    def exportMaintenanceToExcel(self):
        self.cursor.execute("SELECT * FROM maintenance")
        rows = self.cursor.fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ТО"
        ws.append(["ID", "ID автобуса", "Дата", "Тип"])
        for r in rows:
            ws.append(r)
        filename = "maintenance_export.xlsx"
        wb.save(filename)
        QMessageBox.information(self, "Экспорт", f"Данные ТО экспортированы в {filename}")
        self.log_action("Экспорт данных ТО в Excel")

    # --- Проверка уведомлений ТО ---
    def check_maintenance_notifications(self):
        today = datetime.datetime.now().date()
        notify = []
        self.cursor.execute("SELECT bus_id, date, type FROM maintenance")
        rows = self.cursor.fetchall()
        for bus_id, date_str, mtype in rows:
            date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            # Если ТО в ближайшие 7 дней или просрочено
            if 0 <= (date_obj - today).days <= 7 or (date_obj - today).days < 0:
                notify.append(f"Автобус ID {bus_id}: {mtype} назначено на {date_str}")

        if notify:
            QMessageBox.information(self, "Уведомления ТО", "\n".join(notify))

    # --- Генерация отчёта по ТО ---
    def generateMaintenanceReport(self):
        date = self.report_date.date().toString("yyyy-MM-dd")
        self.cursor.execute("SELECT * FROM maintenance WHERE date=?", (date,))
        rows = self.cursor.fetchall()

        self.report_table.setColumnCount(4)
        self.report_table.setRowCount(len(rows))
        self.report_table.setHorizontalHeaderLabels(["ID", "ID автобуса", "Дата", "Тип"])

        for r_idx, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                self.report_table.setItem(r_idx, c_idx, QTableWidgetItem(str(val)))

        self.log_action(f"Сгенерирован отчёт ТО на дату {date}")

def main():
    app = QApplication(sys.argv)

    def start_app(user_id, user_role):
        window = BusParkApp(user_id, user_role)
        window.show()
        app.exec_()

    login = LoginWindow(start_app)
    login.show()

    sys.exit(app.exec_())

if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = None

    def on_login_success(user_id, role):
        global main_window
        main_window = BusParkApp(user_id, role)
        main_window.show()

    login_window = LoginWindow(on_login_success)
    login_window.show()

    sys.exit(app.exec_())

